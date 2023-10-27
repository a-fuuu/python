import os
import shutil
import openpyxl
import requests
import re
import json
from konlpy.tag import Okt

# tag 요청 리스트 상대경로
excel_file_path = "./Tag_library_추가요청.xlsx"

# 파일 다운로드

def file_download():
    path = "/mnt/c/Users/전준표/OneDrive - 하이퍼라운지/Shared Documents/General/☀︎ 팀 회의 자료/DE_데이터팀/400. Tag Library/Tag_library_추가요청.xlsx"

    filename = os.path.basename(path)

    local_path = "./"

    # 최신 tag library 요청 리스트를 받기 위해 파일이 존재하는 경우 삭제
    if os.path.exists("./Tag_library_추가요청.xlsx"):
        os.remove("./Tag_library_추가요청.xlsx")
    
    # 파일을 local path에 복사합니다. Onedrive 경로의 파일은 접근 권한 이슈가 있음
    shutil.copyfile(path, os.path.join(local_path, filename))

def translate(text):
    '''
    네이버 파파고 API를 활용한
    한-영 번역을 수행하는 함수입니다.
    '''

    # API key
    with open('./key.txt', 'r') as key_file:
        contents = key_file.read()
        contents = contents.split('\n')
        client_id = contents[0]
        client_secret = contents[1]

    data = {'text' : text,
            'source' : 'ko',
            'target': 'en'}

    url = "https://openapi.naver.com/v1/papago/n2mt"

    header = {"X-Naver-Client-Id":client_id,
            "X-Naver-Client-Secret":client_secret}

    response = requests.post(url, headers=header, data=data)
    rescode = response.status_code

    if(rescode==200):
        send_data = response.json()
        trans_data = (send_data['message']['result']['translatedText'])
        return trans_data.lower()
    else:
        print("Error Code:" , rescode)

def extract_translate_append():
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb['Sheet1']

    file_path = '../../../professional/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json'
    
    with open(file_path, encoding='utf-8') as f:
        data = json.load(f)
    
    version = data["version"]
    version = version.split(".")
    version = [int(x) for x in version]
    version[-1] = version[-1] + 1
    if version[-1] == 100:
        version[-2] = version[-2] + 1
        version[-1] = 0
    version = [str(x) for x in version]
    version[-1] = "{:0<{width}}".format(version[-1], width=2)
    version = ".".join(version)
    data["version"] = version
    
    subject_area = data["hl_standards_tag"][0]['subject_area']
    sub = [x['id'] for x in subject_area]

    tag_kr = []
    for subject_area in sub:
        element = data['hl_standards_tag'][0]['analysis_item'][0]['key_tag'][0][subject_area]
        kr = [kr['kr'] for kr in element]
        tag_kr = tag_kr + kr

    okt = Okt()
    result = []
    for row in sheet.iter_rows(min_row=sheet.max_row - 10, max_row=sheet.max_row, min_col=14, max_col=14):
        for cell in row:
            row_number = cell.row
            if cell.value == '완료': # 요청상태가 완료인 항목만
            # if cell.value is None: # 요청상태가 완료인 항목만
                if sheet.cell(row=row_number, column=8).value is not None: # 요청 list가 비어있지 않은 항목
                    req = sheet.cell(row=row_number, column=8).value
                    req_list = re.split(r"\n|,", req)

                    # 빈 element 제거
                    req_list = [x for x in req_list if x != '']
                    
                    for val in req_list:
                        kr = re.sub(r"\[.*\]", '', val)
                        if kr in tag_kr:
                            continue
                        en = translate(' '.join(okt.morphs(kr.strip()))) # konlpy 활용하여 형태소 분리 후 번역 (번역 퀄리티 높이기 위함)
                        id = re.sub("-| ", '_', en)
                        output = { "id" : str(id), "en" : str(en), "kr" : str(kr) }
                        result.append(output)
    if result == []:
        print('업데이트 사항이 없습니다.')
        return
    else:
        print('업데이트 사항이 있습니다.')
    for item in result:
        data["hl_standards_tag"][0]["analysis_item"][0]['key_tag'][0]['common'].append(item)
    
    if os.path.exists(f'./hl_standards_new_tag_v{version}.json'):
        os.remove(f'./hl_standards_new_tag_v{version}.json')

    with open(f'./hl_standards_new_tag_v{version}.json', 'w', encoding='utf-8') as output:
        json.dump(data, output, indent=2, ensure_ascii=False)
    wb.close()

if __name__=='__main__':
    # 수행요청 list file 다운로드
    file_download()

    # 다운 받은 excel file 기반으로 추가할 list 정리
    extract_translate_append()