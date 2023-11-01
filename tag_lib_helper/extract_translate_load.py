import os
import shutil
import openpyxl
import requests
import re
import json
from konlpy.tag import Okt

def translate(text):
    '''
    네이버 파파고 API를 활용한
    한-영 번역을 수행하는 함수입니다.
    '''

    # API key
    with open('./key.txt', 'r') as key_file:
        contents = key_file.read()
        contents = contents.split('\n')
        client_id = contents[0]
        client_secret = contents[1]

    data = {'text' : text,
            'source' : 'ko',
            'target': 'en'}

    url = "https://openapi.naver.com/v1/papago/n2mt"

    header = {"X-Naver-Client-Id":client_id,
            "X-Naver-Client-Secret":client_secret}

    response = requests.post(url, headers=header, data=data)
    rescode = response.status_code

    if(rescode==200):
        send_data = response.json()
        trans_data = (send_data['message']['result']['translatedText'])
        return trans_data.lower()
    else:
        print("Error Code:" , rescode)

def extract_translate_append():
    # tag 요청 리스트 상대경로
    excel_file_path = "./Tag_library_추가요청.xlsx"

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb['Sheet1']

    file_path = '../../../professional/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json'
    
    with open(file_path, encoding='utf-8') as f:
        data = json.load(f)
    
    version = data["version"]
    version = version.split(".")
    version = [int(x) for x in version]
    version[-1] = version[-1] + 1
    if version[-1] == 100:
        version[-2] = version[-2] + 1
        version[-1] = 0
    version = [str(x) for x in version]
    version[-1] = "{:0<{width}}".format(version[-1], width=2)
    version = ".".join(version)
    data["version"] = version
    
    subject_area = data["hl_standards_tag"][0]['subject_area']
    sub = [x['id'] for x in subject_area]

    tag_kr = []
    for subject_area in sub:
        element = data['hl_standards_tag'][0]['analysis_item'][0]['key_tag'][0][subject_area]
        kr = [kr['kr'] for kr in element]
        tag_kr = tag_kr + kr

    okt = Okt()
    result = []
    for row in sheet.iter_rows(min_row=sheet.max_row - 10, max_row=sheet.max_row, min_col=14, max_col=14):
        for cell in row:
            row_number = cell.row
            if cell.value == '완료': # 요청상태가 완료인 항목만
                if sheet.cell(row=row_number, column=8).value is not None: # 요청 list가 비어있지 않은 항목
                    req = sheet.cell(row=row_number, column=8).value
                    req_list = re.split(r"\n|,", req)

                    # 빈 element 제거
                    req_list = [x for x in req_list if x != '']
                    
                    for val in req_list:
                        kr = re.sub(r"\[.*\]", '', val).strip()
                        if kr in tag_kr:
                            continue
                        en = translate(' '.join(okt.morphs(kr))) # konlpy 활용하여 형태소 분리 후 번역 (번역 퀄리티 향상)
                        id = re.sub("-| ", '_', en)
                        output = { "id" : str(id), "en" : str(en), "kr" : str(kr) }
                        result.append(output)
    if result == []:
        print('업데이트 사항이 없습니다.')
        return
    else:
        print('업데이트 사항이 있습니다.')
    for item in result:
        data["hl_standards_tag"][0]["analysis_item"][0]['key_tag'][0]['common'].append(item)
    
    if os.path.exists(f'./hl_standards_new_tag_v{version}.json'):
        os.remove(f'./hl_standards_new_tag_v{version}.json')

    with open(f'./lib_lib/hl_standards_new_tag_v{version}.json', 'w', encoding='utf-8') as output:
        json.dump(data, output, indent=2, ensure_ascii=False)

    if os.path.exists(f'./hl_standards_new_tag_v{version}.json'):
        os.remove(f'./hl_standards_new_tag_v{version}.json')

    with open(f'./hl_standards_new_tag.json', 'w', encoding='utf-8') as output:
        json.dump(data, output, indent=2, ensure_ascii=False)

    wb.close()

    return version


if __name__=='__main__':

    # 다운 받은 요청 list 기반으로 tag lib 생성
    extract_translate_append()