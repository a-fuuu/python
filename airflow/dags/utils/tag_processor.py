import os
import openpyxl
import requests
import re
import json
import subprocess
from googletrans import Translator
import pickle

def extract_translate_append(**kwargs):
    duplicated_list_file = '/home/junpyo/airflow/duplicated_list.pkl'

    if os.path.exists(duplicated_list_file):
        with open(duplicated_list_file, 'rb') as f:
            existing_duplicated = pickle.load(f)
    else:
        existing_duplicated = set()

    pattern = re.compile(r'[^a-zA-Z\s]')
    translator = Translator()

    ti = kwargs['ti']
    pull_command = 'cd /home/junpyo/airflow/modelmapping/ && git pull'
    subprocess.call(pull_command, shell=True)

    excel_file_path = '/home/junpyo/airflow/Tag_library_추가요청.xlsx'
    json_file_path = '/home/junpyo/airflow/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json'

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb['Sheet1']
    
    with open(json_file_path, encoding='utf-8') as f:
        data = json.load(f)
    
    version = data["version"]
    version = version.split(".")

    major = int(version[0])
    minor = int(version[1])
    patch = int(version[2])

    patch += 1

    if patch == 100:
        minor += 1
        patch = 1

    version = f"{major}.{minor}.{patch:02d}"
    
    subject_area = data["hl_standards_tag"][0]['subject_area']
    sub = [x['id'] for x in subject_area]

    existing_id = []
    element = []
    tag_kr = []
    for subject_area in sub:
        element = data['hl_standards_tag'][0]['analysis_item'][0]['key_tag'][0][subject_area]
        kr = [item['kr'] for item in element]
        existing_id += [item['id'] for item in element]
        tag_kr += kr

    result = []
    kr_list = []
    id_list = []
    duplicated_kr_list = []
    duplicated_id_list = []
    print(sheet.max_row)
    for row in sheet.iter_rows(min_row=690, max_row=sheet.max_row, min_col=14, max_col=14):
        for cell in row:
            row_number = cell.row
            if cell.value == '완료': # 요청상태가 완료인 항목만
                if sheet.cell(row=row_number, column=8).value is not None: # 요청 list가 비어있지 않은 항목
                    req = sheet.cell(row=row_number, column=8).value
                    req_list = re.split(r"\n|,", req)

                    # 빈 element 제거
                    req_list = [x for x in req_list if x != '']
                    
                    for val in req_list:
                        kr = re.sub(r"\[.*\]", '', val).strip()
                        if re.sub(' ', '', kr) in tag_kr:
                            continue
                        en = translator.translate(((kr)), dest = 'en').text
                        en = re.sub(pattern, ' ', en)
                        id = re.sub("-| ", '_', en)
                        if id.lower() in existing_id:
                            print('중복 발생')
                            if kr in existing_duplicated:
                                continue
                            else:
                                existing_duplicated.add(kr)
                            duplicated_kr_list.append(re.sub(' ', '', kr))
                            duplicated_id_list.append(id)
                            continue
                        kr_list.append(re.sub(' ', '', kr))
                        id_list.append(id.lower())
                        output = { "id" : str(id.lower()), "en" : str(en), "kr" : str(re.sub(' ', '', kr)) }
                        result.append(output)
                        print(result)
                        
    with open(duplicated_list_file, 'wb') as f:
        pickle.dump(existing_duplicated, f)
    if result == [] and duplicated_kr_list == []:
        return None
    
    msg1 = ', '.join(kr_list) # Teams webhook로 보낼 한글 tag 명
    msg2 = ', '.join(id_list) # Teams webhook로 보낼 tag id
    msg3 = ', '.join(duplicated_kr_list) # Teams webhook로 보낼 중복 tag 명
    msg4 = ', '.join(duplicated_id_list) # Teams webhook로 보낼 중복 tag id
    
    # 메시지
    if result == [] and duplicated_kr_list != []:
        message = f"추가되지 않은 Tag {msg3} | 다음의 Tag_id 중복으로 추가되지 않았습니다 -> [{msg4}] Tag명 변경하여 재요청 해주세요"

    elif result != [] and duplicated_kr_list != []:
        message = f"""
        추가된 Tag : {msg1} | Tag_id : {msg2} | 버전 : {version} | modelmapping git pull을 통해 최신 tag library를 받으세요
        추가되지 않은 Tag {msg3} | 다음의 Tag_id 중복으로 추가되지 않았습니다 -> [{msg4}] Tag명 변경하여 재요청 해주세요
        """
        data["version"] = version

    elif result != [] and duplicated_kr_list == []:
        message = f"추가된 Tag : {msg1} | Tag_id : {msg2} | 버전 : {version} | modelmapping git pull을 통해 최신 tag library를 받으세요"
        data["version"] = version

    
    for item in result:
        data["hl_standards_tag"][0]["analysis_item"][0]['key_tag'][0]['common'].append(item)
    
    if os.path.exists(f'/home/junpyo/airflow/modelmapping/de_server/tag_lib_auto/lib_lib/hl_standards_new_tag_v{version}.json'):
        os.remove(f'/home/junpyo/airflow/modelmapping/de_server/tag_lib_auto/lib_lib/hl_standards_new_tag_v{version}.json')

    with open(f'/home/junpyo/airflow/modelmapping/de_server/tag_lib_auto/lib_lib/hl_standards_new_tag_v{version}.json', 'w', encoding='utf-8') as output:
        json.dump(data, output, indent=2, ensure_ascii=False)

    if os.path.exists(f'/home/junpyo/airflow/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json'):
        os.remove(f'/home/junpyo/airflow/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json')
    
    with open(f'/home/junpyo/airflow/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json', 'w', encoding='utf-8') as output:
        json.dump(data, output, indent=2, ensure_ascii=False)

    wb.close()
    
    # Webhook URL
    webhook_url = "https://hyperlounge.webhook.office.com/webhookb2/7f29f864-e756-41af-8045-451d0f42a098@0cad3bb2-0c3d-4882-aa6b-714ad34b84eb/IncomingWebhook/47ead61ee03647b38d034bafd69886c6/6c36f60c-61bc-481d-beac-4d34f81ab9d6"

    
    # Webhook에 메시지를 보냅니다.
    requests.post(webhook_url, json={"text": message})
    
    ti.xcom_push(key='version', value=version)
    return version is not None
