from airflow import DAG
from datetime import datetime, timedelta
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
from airflow.utils.dates import days_ago
import os
import shutil
import openpyxl
import requests
import re
import json
from konlpy.tag import Okt
import subprocess

# from tasks.tag_request_file_download import file_download
# from tasks.extract_translate_load import translate
# from tasks.extract_translate_load import extract_translate_append
# from tasks.git_sync import git_pull_push

def file_download():
    path = "/mnt/c/Users/전준표/OneDrive - 하이퍼라운지/Shared Documents/General/☀︎ 팀 회의 자료/DE_데이터팀/400. Tag Library/Tag_library_추가요청.xlsx"

    filename = os.path.basename(path)

    local_path = "./"

    # 최신 tag library 요청 리스트를 받기 위해 파일이 존재하는 경우 삭제
    if os.path.exists("./Tag_library_추가요청.xlsx"):
        os.remove("./Tag_library_추가요청.xlsx")
    
    # 파일을 local path에 복사합니다. Onedrive 경로의 파일은 접근 권한 이슈가 있음
    shutil.copyfile(path, os.path.join(local_path, filename))

def translate(text):
    '''
    네이버 파파고 API를 활용한
    한-영 번역을 수행하는 함수입니다.
    '''

    # API key
    with open('./key.txt', 'r') as key_file:
        contents = key_file.read()
        contents = contents.split('\n')
        client_id = contents[0]
        client_secret = contents[1]

    data = {'text' : text,
            'source' : 'ko',
            'target': 'en'}

    url = "https://openapi.naver.com/v1/papago/n2mt"

    header = {"X-Naver-Client-Id":client_id,
            "X-Naver-Client-Secret":client_secret}

    response = requests.post(url, headers=header, data=data)
    rescode = response.status_code

    if(rescode==200):
        send_data = response.json()
        trans_data = (send_data['message']['result']['translatedText'])
        return trans_data.lower()
    else:
        print("Error Code:" , rescode)

def extract_translate_append(**kwargs):
    ti = kwargs['ti']
    pull_command = 'cd /home/junpyo/airflow/modelmapping/ && git pull'
    subprocess.call(pull_command, shell=True)

    # tag 요청 리스트 상대경로
    excel_file_path = '/home/junpyo/airflow/Tag_library_추가요청.xlsx'

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb['Sheet1']

    file_path = '/home/junpyo/airflow/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json'
    
    with open(file_path, encoding='utf-8') as f:
        data = json.load(f)
    
    version = data["version"]
    version = version.split(".")
    version = [int(x) for x in version]
    version[-1] = version[-1] + 1
    if version[-1] == 100:
        version[-2] = version[-2] + 1
        version[-1] = 0
    version = [str(x) for x in version]
    version[-1] = "{:0<{width}}".format(version[-1], width=2)
    version = ".".join(version)
    data["version"] = version
    
    subject_area = data["hl_standards_tag"][0]['subject_area']
    sub = [x['id'] for x in subject_area]

    tag_kr = []
    for subject_area in sub:
        element = data['hl_standards_tag'][0]['analysis_item'][0]['key_tag'][0][subject_area]
        kr = [kr['kr'] for kr in element]
        tag_kr = tag_kr + kr

    okt = Okt()
    result = []
    for row in sheet.iter_rows(min_row=sheet.max_row - 10, max_row=sheet.max_row, min_col=14, max_col=14):
        print(sheet.max_row)
        for cell in row:
            row_number = cell.row
            if cell.value == '완료': # 요청상태가 완료인 항목만
                if sheet.cell(row=row_number, column=8).value is not None: # 요청 list가 비어있지 않은 항목
                    req = sheet.cell(row=row_number, column=8).value
                    req_list = re.split(r"\n|,", req)

                    # 빈 element 제거
                    req_list = [x for x in req_list if x != '']
                    
                    for val in req_list:
                        kr = re.sub(r"\[.*\]", '', val).strip()
                        print(kr)
                        if kr in tag_kr:
                            continue
                        en = translate(' '.join(okt.morphs(kr))) # konlpy 활용하여 형태소 분리 후 번역 (번역 퀄리티 향상)
                        id = re.sub("-| ", '_', en)
                        output = { "id" : str(id), "en" : str(en), "kr" : str(kr) }
                        result.append(output)
    if result == []:
        return None

    for item in result:
        data["hl_standards_tag"][0]["analysis_item"][0]['key_tag'][0]['common'].append(item)
    
    if os.path.exists(f'/home/junpyo/airflow/modelmapping/de_server/tag_lib_auto/lib_lib/hl_standards_new_tag_v{version}.json'):
        os.remove(f'/home/junpyo/airflow/modelmapping/de_server/tag_lib_auto/lib_lib/hl_standards_new_tag_v{version}.json')

    # if not os.path.exists(f'/home/junpyo/airflow/modelmapping/de_server/tag_lib_auto/lib_lib/test/'):
    #     os.makedirs(f'/home/junpyo/airflow/modelmapping/de_server/tag_lib_auto/lib_lib/test/')

    with open(f'/home/junpyo/airflow/modelmapping/de_server/tag_lib_auto/lib_lib/hl_standards_new_tag_v{version}.json', 'w', encoding='utf-8') as output:
        json.dump(data, output, indent=2, ensure_ascii=False)

    if os.path.exists(f'/home/junpyo/airflow/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json'):
        os.remove(f'/home/junpyo/airflow/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json')

    # if not os.path.exists(f'/home/junpyo/airflow/modelmapping/de_server/db_to_excel/test/'):
    #     os.makedirs(f'/home/junpyo/airflow/modelmapping/de_server/db_to_excel/test/')
    
    with open(f'/home/junpyo/airflow/modelmapping/de_server/db_to_excel/hl_standards_new_tag.json', 'w', encoding='utf-8') as output:
        json.dump(data, output, indent=2, ensure_ascii=False)

    wb.close()
    print(version)
    ti.xcom_push(key='version', value=version)
    return version

def git_pull_push(**kwargs):
    # version = kwargs['ti'].xcom_pull(task_ids='extract_translate_append')
    # print(version)
    # if version is None:
    #     print('업데이트 사항이 없습니다.')
    # else:
        # print(f'업데이트 사항이 있습니다. version : {version}')

    change_directory = 'cd /home/junpyo/airflow/modelmapping/ && git pull'
    subprocess.call(change_directory, shell=True)

    # Define the working directory
    working_directory = "/home/junpyo/airflow/modelmapping/"
    
    add_command = "git add ."
    subprocess.run(add_command, shell=True, cwd=working_directory)

    commit_command = 'git commit -m "tag lib update"'
    subprocess.call(commit_command, shell=True, cwd=working_directory)

    push_command = 'git push'
    subprocess.call(push_command, shell=True, cwd=working_directory)

default_args = {
    'start_date': days_ago(1),
}

dag = DAG(
    dag_id='Tag_Library_update',
    default_args=default_args,
    schedule_interval="*/10 * * * *",  # This sets the schedule to run every 10 minutes
)

download = PythonOperator(
    task_id='file_download',
    python_callable= file_download,
    dag=dag
)

etl = PythonOperator(
    task_id='etl',
    python_callable=extract_translate_append,
    provide_context=True,
    dag=dag
)

git_sync = PythonOperator(
    task_id='git_sync',
    python_callable=git_pull_push,
    provide_context=True,
    dag=dag
    )

download >> etl >> git_sync